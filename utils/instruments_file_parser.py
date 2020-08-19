import config, logging, os, traceback
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(config.logger_name)


class LondonStockExchangeFileParser():

    def __init__(self):

        if os.path.isfile(config.lse_instruments_filename):

            self.wb = load_workbook(filename=config.lse_instruments_filename)

            min_row = 0
            min_col = 1
            max_row = 0

            ws = self.wb[config.lse_instruments_main_sheetname]
            row_count = ws.max_row
            max_col = ws.max_column

            # Calculating max_row

            if ws.cell(row=row_count-1, column=1).value is None:

                logger.warning('LondonStockExchangeFileParser: getting new row_count')

                for i in range(row_count-5, row_count-2):

                    cell = ws.cell(row=i, column=1)
                    if cell.value is not None:
                        max_row = i

            else:
                max_row = row_count


            # Calculating min row

            if ws.cell(row=8, column=1).value != "TIDM":

                logger.debug('TIDM not found at A8' + str(ws.cell(row=8, column=1).value))

                for i in range(1, 20):
                    if ws.cell(row=i, column=1).value == "TIDM":
                        min_row = i
                        break
            else:
                min_row = 8


            # Read the cell values into a list of lists
            data_rows = []
            for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
                data_cols = []
                for cell in row:
                    data_cols.append(cell.value)
                data_rows.append(data_cols)

            # Transform into dataframe
            df = pd.DataFrame(data_rows)
            new_header = df.iloc[0]
            df = df[1:]
            df.columns = new_header

            self.df = df
            self.df['exchange']='LSE'

        else:
            logger.error('LondonStockExchangeFileParser#__init__: lse instruments file does not exist')
            logger.debug(traceback.format_exc())

    def get_raw(self):

        return self.df

    def get_database_data(self):
        self.df['ex_symbol'] = self.df['exchange'] + self.df['TIDM']
        self.df.rename(columns={'TIDM':'symbol',
                           'Issuer Name':'name',
                           'LSE Market':'market'},
                  inplace=True)
        return self.df[['ex_symbol', 'symbol', 'exchange', 'market', 'name']]


    def get_all_tickers(self):

        return self.df['TIDM']

    def get_aim_market(self):

        return self.df.loc[self.df['LSE Market'] == "AIM"]

    def get_main_market(self):

        return self.df.loc[self.df['LSE Market'] == "MAIN MARKET"]

    def get_main_market_hgs(self):

        return self.df.loc[self.df['LSE Market'] == "MAIN MARKET - HGS"]

    def get_main_market_sfs(self):

        return self.df.loc[self.df['LSE Market'] == "MAIN MARKET - SFS"]

    def get_professional_securities_market(self):

        return self.df.loc[self.df['LSE Market'] == "PROFESSIONAL SECURITIES MARKET"]

    def get_trading_only_market(self):

        return self.df.loc[self.df['LSE Market'] == "ADMISSION TO TRADING ONLY"]


class NASDAQFileParser():

    def __init__(self):

        if os.path.isfile(config.nasdaq_instruments_filename):

            df = pd.read_csv(config.nasdaq_instruments_filename, sep="|", header=0, skipfooter=1, engine='python')

            self.df = df
            self.df['exchange']='NASDAQ'


        else:
            logger.error('NASDAQFileParser#__init__: nasdaq instruments file does not exist')
            logger.debug(traceback.format_exc())

    def get_raw(self):

        return self.df

    def get_database_data(self):

        self.df['ex_symbol'] = self.df['exchange'] + self.df['Symbol']

        # Remove Test Issue Instruments
        self.df = self.df[self.df['Test Issue'] != 'Y']

        # Parse name to remove instruments type (split on '-')
        self.df['Security Name'] = self.df['Security Name'].str.split('-').str[0]

        # Q = NASDAQ Global Select MarketSM
        # G = NASDAQ Global MarketSM
        # S = NASDAQ Capital Market
        self.df['Market Category'] = self.df['Market Category'].replace({"Q": "Global Select MarketSM",
                                                                         "G": "Global MarketSM",
                                                                         "S": "Capital Market"
                                                                         })
        self.df.rename(columns={'Symbol': 'symbol',
                                'Security Name': 'name',
                                'Market Category': 'market'},
                       inplace=True)

        return self.df[['ex_symbol', 'symbol', 'exchange', 'market', 'name']]


class NYSEFileParser():

    def __init__(self):

        if os.path.isfile(config.nyse_instruments_filename):

            df = pd.read_csv(config.nyse_instruments_filename, sep="|", header=0, skipfooter=1, engine='python')

            self.df = df
            self.df['exchange']='NYSE'

            # Exchanges
            # A = NYSE MKT
            # N = New York Stock Exchange (NYSE)
            # P = NYSE ARCA
            # Z = BATS Global Markets (BATS)
            # V = Investors' Exchange, LLC (IEXG)

            # Removing BATS and IEXG
            self.df = self.df[self.df['Exchange'] != 'Z']
            self.df = self.df[self.df['Exchange'] != 'V']


        else:
            logger.error('NYSEFileParser#__init__: nyse instruments file does not exist')
            logger.debug(traceback.format_exc())

    def get_raw(self):

        return self.df

    def get_database_data(self):

        self.df['ex_symbol'] = self.df['exchange'] + self.df['ACT Symbol']

        # Remove Test Issue Instruments
        self.df = self.df[self.df['Test Issue'] != 'Y']

        # A = NYSE MKT
        # N = New York Stock Exchange (NYSE)
        # P = NYSE ARCA
        self.df['Exchange'] = self.df['Exchange'].replace({"A": "NYSE MKT",
                                                                         "N": "New York Stock Exchange (NYSE)",
                                                                         "P": "NYSE ARCA"
                                                                         })
        self.df.rename(columns={'ACT Symbol': 'symbol',
                                'Security Name': 'name',
                                'Exchange': 'market'},
                       inplace=True)

        return self.df[['ex_symbol', 'symbol', 'exchange', 'market', 'name']]
